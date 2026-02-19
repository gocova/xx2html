import tempfile
import unittest
from pathlib import Path

from bs4 import BeautifulSoup, Comment
from openpyxl import Workbook

import xx2html.core as core_module
from xx2html import create_xlsx_transform


SHEET_HTML = (
    '<section id="{enc_sheet_name}" data-sheet="{sheet_name}">'
    "{table_generated_html}"
    "</section>"
)
SHEETNAME_HTML = '<a class="sheet-nav" href="#{enc_sheet_name}">{sheet_name}</a>'
INDEX_HTML = (
    "<!doctype html><html><head>"
    "{fonts_html}{core_css_html}{user_css_html}{generated_css_html}"
    "{generated_incell_css_html}{conditional_css_html}"
    "</head><body data-source=\"{source_filename}\">{sheets_names_generated_html}{sheets_generated_html}{safari_js}</body></html>"
)


def _build_transform(**kwargs):
    return create_xlsx_transform(
        sheet_html=SHEET_HTML,
        sheetname_html=SHEETNAME_HTML,
        index_html=INDEX_HTML,
        fonts_html="",
        core_css="",
        user_css="",
        safari_js="",
        apply_cf=False,
        **kwargs,
    )


class TransformOptionsTests(unittest.TestCase):
    def test_create_transform_validates_required_placeholders(self):
        with self.assertRaises(ValueError) as error_context:
            create_xlsx_transform(
                sheet_html="<section>{sheet_name}</section>",
                sheetname_html=SHEETNAME_HTML,
                index_html=INDEX_HTML,
                fonts_html="",
                core_css="",
                user_css="",
                safari_js="",
            )
        self.assertIn("sheet_html", str(error_context.exception))
        self.assertIn("table_generated_html", str(error_context.exception))

    def test_create_transform_validates_preview_limits(self):
        with self.assertRaises(ValueError):
            _build_transform(max_sheets=0)
        with self.assertRaises(ValueError):
            _build_transform(max_rows=0)
        with self.assertRaises(ValueError):
            _build_transform(max_cols=0)

    def test_raise_on_error_propagates_original_exception(self):
        transform = _build_transform(raise_on_error=True)
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_file = Path(tmp_dir) / "output.html"
            with self.assertRaises(FileNotFoundError):
                transform("__missing__/missing.xlsx", str(output_file), "en_US")

    def test_failed_transform_does_not_truncate_existing_destination_file(self):
        transform = _build_transform()
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_file = Path(tmp_dir) / "output.html"
            output_file.write_text("existing-content", encoding="utf-8")

            ok, err = transform("__missing__/missing.xlsx", str(output_file), "en_US")

            self.assertFalse(ok)
            self.assertIsInstance(err, str)
            self.assertEqual("existing-content", output_file.read_text(encoding="utf-8"))

    def test_max_sheets_limits_visible_sheets(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_file = Path(tmp_dir) / "source.xlsx"
            output_file = Path(tmp_dir) / "output.html"

            workbook = Workbook()
            first_sheet = workbook.active
            first_sheet.title = "First"
            first_sheet["A1"] = "First value"
            second_sheet = workbook.create_sheet("Second")
            second_sheet["A1"] = "Second value"
            workbook.save(source_file)
            workbook.close()

            transform = _build_transform(max_sheets=1)
            ok, err = transform(str(source_file), str(output_file), "en_US")

            self.assertTrue(ok, err)
            html = output_file.read_text(encoding="utf-8")
            self.assertIn('data-sheet="First"', html)
            self.assertNotIn('data-sheet="Second"', html)

    def test_max_rows_and_cols_limit_rendered_cells(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_file = Path(tmp_dir) / "source.xlsx"
            output_file = Path(tmp_dir) / "output.html"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            for row_index in range(1, 4):
                for col_index in range(1, 4):
                    worksheet.cell(
                        row=row_index, column=col_index, value=f"R{row_index}C{col_index}"
                    )
            workbook.save(source_file)
            workbook.close()

            transform = _build_transform(max_rows=2, max_cols=2)
            ok, err = transform(str(source_file), str(output_file), "en_US")

            self.assertTrue(ok, err)
            html = output_file.read_text(encoding="utf-8")
            self.assertIn('id="Data!A1"', html)
            self.assertIn('id="Data!B2"', html)
            self.assertNotIn('id="Data!C1"', html)
            self.assertNotIn('id="Data!A3"', html)

    def test_generated_html_includes_generator_meta_and_comment(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            source_file = Path(tmp_dir) / "source.xlsx"
            output_file = Path(tmp_dir) / "output.html"

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Data"
            worksheet["A1"] = "hello"
            workbook.save(source_file)
            workbook.close()

            transform = _build_transform()
            ok, err = transform(str(source_file), str(output_file), "en_US")

            self.assertTrue(ok, err)
            html = output_file.read_text(encoding="utf-8")
            soup = BeautifulSoup(html, "lxml")

            generator_meta = soup.find("meta", attrs={"name": "generator"})
            self.assertIsNotNone(generator_meta)
            self.assertEqual(
                f"xx2html {core_module._get_xx2html_version()}",
                generator_meta.get("content"),
            )

            body = soup.body
            self.assertIsNotNone(body)
            body_comments = [
                str(comment).strip()
                for comment in body.find_all(string=lambda value: isinstance(value, Comment))
            ]
            self.assertIn(
                f"Generated by xx2html {core_module._get_xx2html_version()}",
                body_comments,
            )


if __name__ == "__main__":
    unittest.main()
