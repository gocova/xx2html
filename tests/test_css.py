import unittest

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from xx2html.core.css import (
    CssRegistry,
    create_get_css_components_from_cell,
    get_border_classes_from_cell,
)


class CssRegistryTests(unittest.TestCase):
    @staticmethod
    def _color_resolver(color):
        rgb = getattr(color, "rgb", None)
        return rgb if isinstance(rgb, str) else None

    def test_register_font_underline_uses_text_decoration(self):
        registry = CssRegistry(get_css_color=lambda _: None)

        class_name = registry.register_font_underline()

        self.assertEqual("xlsx_cell_font_underline", class_name)
        self.assertEqual("text-decoration: underline;", registry.classes[class_name])

    def test_default_classes_dict_is_not_shared_between_instances(self):
        registry_a = CssRegistry(get_css_color=lambda _: None)
        registry_b = CssRegistry(get_css_color=lambda _: None)

        registry_a.register_font_bold()

        self.assertIn("xlsx_cell_font_bold", registry_a.classes)
        self.assertNotIn("xlsx_cell_font_bold", registry_b.classes)

    def test_register_core_class_helpers(self):
        registry = CssRegistry(get_css_color=self._color_resolver)

        self.assertEqual("xlsx_cell_font_size_12", registry.register_font_size(12))
        self.assertEqual("xlsx_element_height_33", registry.register_height(33))
        self.assertEqual(
            "xlsx_cell_font_bold", registry.register_font_bold()
        )
        self.assertEqual(
            "xlsx_cell_font_italic", registry.register_font_italic()
        )
        self.assertEqual(
            "xlsx_cell_font_color_FF112233",
            registry.register_font_color(Side(color="FF112233").color),
        )
        self.assertEqual(
            "xlsx_cell_background_color_FF010203",
            registry.register_background_color(Side(color="FF010203").color),
        )

        self.assertIn("font-size: 12px;", registry.classes.values())
        self.assertIn("height: 33px;", registry.classes.values())
        self.assertIn("font-weight: bold;", registry.classes.values())
        self.assertIn("font-style: italic;", registry.classes.values())

    def test_register_color_methods_return_none_when_color_unresolved(self):
        registry = CssRegistry(get_css_color=lambda _: None)
        color = Side(color="FF112233").color

        self.assertIsNone(registry.register_font_color(color))
        self.assertIsNone(registry.register_background_color(color))

    def test_register_border_uses_default_style_and_optional_color(self):
        registry = CssRegistry(get_css_color=self._color_resolver)
        color = Side(color="FFAABBCC").color

        class_name = registry.register_border("dashDot", "left", color)

        self.assertEqual("border_dashDot_l_FFAABBCC", class_name)
        class_value = registry.classes[class_name]
        self.assertIn("border-left-style: solid;", class_value)
        self.assertIn("border-left-width: 1px;", class_value)
        self.assertIn("border-left-color:", class_value)

    def test_register_border_returns_none_when_style_none(self):
        registry = CssRegistry(get_css_color=self._color_resolver)
        color = Side(color="FF000000").color
        self.assertIsNone(registry.register_border(None, "top", color))

    def test_get_border_classes_from_cell_reads_all_sides(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.border = Border(
            left=Side(style="thin", color="FF0000FF"),
            right=Side(style="dashed", color="FF00FF00"),
            top=Side(style="double", color="FFFF0000"),
            bottom=Side(style="medium", color="FF123456"),
        )
        registry = CssRegistry(get_css_color=self._color_resolver)

        classes = get_border_classes_from_cell(cell, registry)

        self.assertEqual(4, len(classes))
        self.assertTrue(any(cls.startswith("border_thin_l_") for cls in classes))
        self.assertTrue(any(cls.startswith("border_dashed_r_") for cls in classes))
        self.assertTrue(any(cls.startswith("border_double_t_") for cls in classes))
        self.assertTrue(any(cls.startswith("border_medium_b_") for cls in classes))

    def test_get_css_components_from_cell_collects_alignment_fill_font_and_merged_borders(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        merged_neighbor = ws["B1"]

        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
        cell.font = Font(
            sz=14,
            color="FF112233",
            b=True,
            i=True,
            u="single",
        )
        cell.border = Border(left=Side(style="thin", color="FF0000FF"))
        merged_neighbor.border = Border(right=Side(style="thick", color="FF101010"))

        registry = CssRegistry(get_css_color=self._color_resolver)
        get_components = create_get_css_components_from_cell(registry)

        styles, classes = get_components(cell, merged_cell_map={"cells": [merged_neighbor]})

        self.assertEqual("center", styles["text-align"])
        self.assertEqual("top", styles["vertical-align"])
        self.assertIn("xlsx_cell_font_size_14", classes)
        self.assertIn("xlsx_cell_font_bold", classes)
        self.assertIn("xlsx_cell_font_italic", classes)
        self.assertIn("xlsx_cell_font_underline", classes)
        self.assertIn("xlsx_cell_font_color_FF112233", classes)
        self.assertIn("xlsx_cell_background_color_FF00FF00", classes)
        self.assertTrue(any(cls.startswith("border_thin_l_") for cls in classes))
        self.assertTrue(any(cls.startswith("border_thick_r_") for cls in classes))

    def test_get_css_components_from_cell_warns_on_unsupported_pattern_type(self):
        wb = Workbook()
        ws = wb.active
        cell = ws["A1"]
        cell.fill = PatternFill(fill_type="darkGrid", fgColor="FF00FF00")

        registry = CssRegistry(get_css_color=self._color_resolver)
        get_components = create_get_css_components_from_cell(registry)

        with self.assertLogs(level="WARNING") as log_ctx:
            styles, classes = get_components(cell)

        self.assertEqual({}, styles)
        self.assertIsInstance(classes, set)
        self.assertTrue(
            any("Pattern type is not supported: darkGrid" in msg for msg in log_ctx.output)
        )


if __name__ == "__main__":
    unittest.main()
