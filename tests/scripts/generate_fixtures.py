from __future__ import annotations

from pathlib import Path
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from PIL import Image

ROOT_DIR = Path(__file__).resolve().parents[2]
FIXTURES_DIR = ROOT_DIR / "tests" / "fixtures"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
RD_NS = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"


def _ensure_content_type_entries(content_types_path: Path) -> None:
    ET.register_namespace("", CONTENT_TYPES_NS)
    tree = ET.parse(content_types_path)
    root = tree.getroot()

    def has_default(ext: str) -> bool:
        return root.find(
            f"{{{CONTENT_TYPES_NS}}}Default[@Extension='{ext}']"
        ) is not None

    def has_override(part_name: str) -> bool:
        return root.find(
            f"{{{CONTENT_TYPES_NS}}}Override[@PartName='{part_name}']"
        ) is not None

    if not has_default("png"):
        root.append(
            ET.Element(
                f"{{{CONTENT_TYPES_NS}}}Default",
                {"Extension": "png", "ContentType": "image/png"},
            )
        )

    overrides = {
        "/xl/metadata.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheetMetadata+xml",
        "/xl/richData/rdrichvalue.xml": "application/vnd.ms-excel.rdrichvalue+xml",
        "/xl/richData/rdrichvaluestructure.xml": "application/vnd.ms-excel.rdrichvaluestructure+xml",
        "/xl/richData/richValueRel.xml": "application/vnd.ms-excel.richvaluerel+xml",
    }
    for part_name, content_type in overrides.items():
        if has_override(part_name):
            continue
        root.append(
            ET.Element(
                f"{{{CONTENT_TYPES_NS}}}Override",
                {"PartName": part_name, "ContentType": content_type},
            )
        )

    tree.write(content_types_path, encoding="utf-8", xml_declaration=True)


def _set_vm_attribute_on_a1(sheet_path: Path) -> None:
    ET.register_namespace("", MAIN_NS)
    tree = ET.parse(sheet_path)
    root = tree.getroot()
    sheet_data = root.find(f"{{{MAIN_NS}}}sheetData")
    if sheet_data is None:
        raise RuntimeError("sheetData not found in sheet XML")

    a1_cell = root.find(f".//{{{MAIN_NS}}}c[@r='A1']")
    if a1_cell is None:
        raise RuntimeError("A1 cell not found in sheet XML")

    a1_cell.set("vm", "1")
    tree.write(sheet_path, encoding="utf-8", xml_declaration=True)


def _write_incell_rich_data_parts(tmp_dir: Path) -> None:
    rich_data_dir = tmp_dir / "xl" / "richData"
    rich_data_rels_dir = rich_data_dir / "_rels"
    media_dir = tmp_dir / "xl" / "media"

    rich_data_dir.mkdir(parents=True, exist_ok=True)
    rich_data_rels_dir.mkdir(parents=True, exist_ok=True)
    media_dir.mkdir(parents=True, exist_ok=True)

    metadata_xml = f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<metadata xmlns=\"{MAIN_NS}\">
  <valueMetadata>
    <bk>
      <rc v=\"0\"/>
    </bk>
  </valueMetadata>
</metadata>
"""
    (tmp_dir / "xl" / "metadata.xml").write_text(metadata_xml, encoding="utf-8")

    rich_structure_xml = f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<rd:richValueStructures xmlns:rd=\"{RD_NS}\">
  <rd:s t=\"_localImage\"/>
</rd:richValueStructures>
"""
    (rich_data_dir / "rdrichvaluestructure.xml").write_text(
        rich_structure_xml, encoding="utf-8"
    )

    rich_value_xml = f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<rd:richValueData xmlns:rd=\"{RD_NS}\">
  <rd:rv s=\"0\"><rd:v>0</rd:v><rd:v>inline-image</rd:v></rd:rv>
</rd:richValueData>
"""
    (rich_data_dir / "rdrichvalue.xml").write_text(rich_value_xml, encoding="utf-8")

    rels_xml = f"""<?xml version=\"1.0\" encoding=\"UTF-8\"?>
<Relationships xmlns=\"{PKG_REL_NS}\">
  <Relationship Id=\"rId1\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/image\" Target=\"../media/image1.png\"/>
</Relationships>
"""
    (rich_data_rels_dir / "richValueRel.xml.rels").write_text(rels_xml, encoding="utf-8")

    image = Image.new("RGBA", (140, 70), (255, 0, 0, 255))
    image.save(media_dir / "image1.png", format="PNG")


def _zip_dir_to_xlsx(source_dir: Path, output_path: Path) -> None:
    with ZipFile(output_path, "w", compression=ZIP_DEFLATED) as archive:
        for file_path in sorted(source_dir.rglob("*")):
            if file_path.is_dir():
                continue
            archive.write(file_path, file_path.relative_to(source_dir).as_posix())


def create_merged_cells_cf_fixture(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    ws["A1"] = "Merged"
    ws.merge_cells("A1:B2")

    ws["C1"] = 10
    ws["C2"] = 2

    warning_fill = PatternFill(start_color="FFFFAA00", end_color="FFFFAA00", fill_type="solid")
    ws.conditional_formatting.add(
        "C1:C2",
        FormulaRule(formula=["TRUE"], fill=warning_fill),
    )

    wb.save(output_path)


def create_incell_image_fixture(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Images"
    ws["A1"] = "InCell"
    ws.column_dimensions["A"].width = 24
    ws.row_dimensions[1].height = 42
    wb.save(output_path)

    with tempfile.TemporaryDirectory() as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)

        with ZipFile(output_path, "r") as archive:
            archive.extractall(tmp_dir)

        _set_vm_attribute_on_a1(tmp_dir / "xl" / "worksheets" / "sheet1.xml")
        _write_incell_rich_data_parts(tmp_dir)
        _ensure_content_type_entries(tmp_dir / "[Content_Types].xml")

        _zip_dir_to_xlsx(tmp_dir, output_path)


def main() -> None:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    create_merged_cells_cf_fixture(FIXTURES_DIR / "merged_cells_cf.xlsx")
    create_incell_image_fixture(FIXTURES_DIR / "incell_image.xlsx")
    print("Generated fixtures in", FIXTURES_DIR)


if __name__ == "__main__":
    main()
