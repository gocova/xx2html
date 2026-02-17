from typing import Callable, List, Set, Tuple

import logging
from zipfile import ZipFile

from openpyxl import load_workbook
from openpyxl.styles.differential import DifferentialStyleList

# Monkey patch!
from xx2html.core.cf import apply_cf_styles
from xx2html.core.patches.openpyxl import apply_patches

from .incell import get_incell_css
from .links import update_links
from .utils import cova_render_table, get_worksheet_contents
from .vm import get_incell_images_refs

# from .css import CssRegistry, create_get_css_components_from_cell
from condif2css.processor import process_conditional_formatting
from condif2css.themes import get_theme_colors
from condif2css.core import create_themed_css_color_resolver
from condif2css.color import argb_to_css
from condif2css.css import CssBuilder, CssRulesRegistry, create_get_css_from_cell

_PATCHES_APPLIED = False


def apply_openpyxl_patches() -> None:
    """Apply required openpyxl monkey patches once per process."""
    global _PATCHES_APPLIED
    if _PATCHES_APPLIED:
        return
    logging.debug("xx2html: applying required openpyxl monkey patches")
    apply_patches()
    _PATCHES_APPLIED = True


# Explicitly apply monkey patches required for vm-aware parsing.
apply_openpyxl_patches()


def create_xlsx_transform(
    sheet_html: str,
    sheetname_html: str,
    index_html: str,
    fonts_html: str,
    core_css: str,
    user_css: str,
    safari_js: str,
    update_local_links: bool = True,
    # prepare_iframe_noscript: bool = True,
    apply_cf: bool = False,
    fail_ok: bool = True,
) -> Callable[[str, str, str], Tuple[bool, None | str]]:
    def transform_xlsx(
        source: str, dest: str, locale: str
    ) -> Tuple[bool, None | str]:  # -> (bool, None|str):
        workbook = None
        workbook_archive = None
        try:
            logging.info(f"Transform (out): Opening '{dest}' for writing...")
            with open(dest, "w", encoding="utf-8") as output:
                sheet_navigation_links = []
                sheet_html_sections = []

                logging.info(f"Transform (wb): Reading '{source}' as xlsx file...")
                workbook = load_workbook(source, data_only=True, rich_text=True)

                logging.debug("Transform (wb|css): Reading theme colors...")
                theme_argb_palette = get_theme_colors(workbook)
                pre_get_css_color = create_themed_css_color_resolver(theme_argb_palette)

                def get_css_color(color):
                    argb_color = pre_get_css_color(color)
                    if argb_color is None:
                        return None
                    return argb_to_css(argb_color)

                css_builder = CssBuilder(get_css_color)
                css_registry = CssRulesRegistry()
                get_css_from_cell = create_get_css_from_cell(
                    css_registry, css_builder=css_builder
                )

                css_cf_registry = CssRulesRegistry(prefix="xx2h_cf")
                get_cf_css_from_diff = create_get_css_from_cell(
                    css_registry=css_cf_registry, css_builder=css_builder
                )

                logging.debug("Transform (wb|incell): Reading incell images...")
                incell_images_refs = {}
                try:
                    workbook_archive = ZipFile(source, "r")
                    incell_images_refs, incell_error = get_incell_images_refs(
                        workbook_archive
                    )
                    if incell_error is not None:
                        raise incell_error
                    logging.info("Transform (wb|incell): Reading complete!")
                except Exception as incell_exc:
                    logging.warning(
                        "Transform (wb|incell): Unable to read incell images due to: %r",
                        incell_exc,
                    )
                    if workbook_archive is not None:
                        workbook_archive.close()
                        workbook_archive = None

                vm_ids = set()
                vm_ids_dimension_references = dict()
                vm_cell_vm_ids = dict()

                encoded_sheet_names = dict()
                sheet_names = workbook.sheetnames
                conditional_formatting_rule_details = {}

                for sheet_name in sheet_names:
                    worksheet = workbook[sheet_name]

                    worksheet_index = workbook.index(worksheet)
                    encoded_sheet_name = f"sheet_{hex(worksheet_index)[2:].zfill(3)}"

                    encoded_sheet_names[sheet_name] = encoded_sheet_name

                    if worksheet.sheet_state == "visible":
                        logging.info(
                            f"Application (ws): Sheet[{worksheet_index}]:'{sheet_name}' (enc_sheet_name: {encoded_sheet_name}) -> is visible"
                        )

                        contents = get_worksheet_contents(
                            worksheet,
                            css_rules_registry=css_registry,
                            css_builder=css_builder,
                            get_css_from_cell=get_css_from_cell,
                            locale=locale,
                            ws_index=worksheet_index,
                        )

                        logging.info(
                            f" {encoded_sheet_name} --> vm_ids: {contents['vm_ids']}"
                        )
                        vm_ids.update(contents["vm_ids"])
                        vm_ids_dimension_references.update(
                            contents["vm_ids_dimension_references"]
                        )
                        vm_cell_vm_ids.update(contents["vm_cell_vm_ids"])

                        sheet_html_sections.append(
                            sheet_html.format(
                                enc_sheet_name=encoded_sheet_name,
                                sheet_name=sheet_name,
                                table_generated_html=cova_render_table(contents),
                            )
                        )

                        sheet_navigation_links.append(
                            sheetname_html.format(
                                enc_sheet_name=encoded_sheet_name, sheet_name=sheet_name
                            )
                        )

                        if apply_cf:
                            logging.info(
                                f"Application (wb|cf): Processing conditional formatting for '{sheet_name}'"
                            )
                            conditional_formatting_rule_details.update(
                                process_conditional_formatting(
                                    worksheet, fail_ok=fail_ok
                                )
                            )

                generated_css = "\n".join(css_registry.get_rules())

                if workbook_archive is not None:
                    logging.debug(
                        "Transform (wb|incell): Preparing incell images output..."
                    )
                    generated_incell_css = get_incell_css(
                        vm_ids,
                        vm_ids_dimension_references,
                        vm_cell_vm_ids,
                        incell_images_refs,
                        workbook_archive,
                    )
                else:
                    generated_incell_css = ""

                logging.info(
                    f"Transform (html|1): Pass 1 --> Preparing {len(conditional_formatting_rule_details)} conditional formatting styles..."
                )
                cf_style_relations: List[Tuple[str, str, Set[str]]] = []
                if hasattr(workbook, "_differential_styles") and isinstance(
                    workbook._differential_styles,  # type: ignore
                    DifferentialStyleList,
                ):
                    for _, details in conditional_formatting_rule_details.items():
                        sheet_name, cell_ref, _, dxf_id, _ = details
                        class_names = get_cf_css_from_diff(
                            workbook._differential_styles[dxf_id],  # type: ignore
                            is_important=True,
                        )
                        cf_style_relations.append((sheet_name, cell_ref, class_names))
                logging.debug(
                    f"Transform: Resulting conditional formatting styles: {cf_style_relations}"
                )

                logging.info("Transform (html|2): Pass 2 --> Preparing html")
                css_rules = "\n".join(css_cf_registry.get_rules())
                html = (
                    index_html.format(
                        sheets_generated_html="\n".join(sheet_html_sections),
                        sheets_names_generated_html="\n".join(sheet_navigation_links),
                        source_filename=source,
                        fonts_html=fonts_html,
                        core_css_html=f"<style>{core_css}</style>",
                        user_css_html=f"<style>{user_css}</style>",
                        generated_css_html=f"<style>{generated_css}</style>",
                        generated_incell_css_html=f"<style>{generated_incell_css}</style>",
                        safari_js=f"<script>{safari_js}</script>",
                        conditional_css_html=f"<style>/*conditional formatting*/\n{css_rules}</style>",
                    )
                    .replace('"$"', "$")
                    .replace('"-"', "-")
                )

                logging.info("Transform (html|3): Pass 3 --> Updating links...")
                html_with_updated_links = update_links(
                    html, encoded_sheet_names, update_local_links=update_local_links
                )

                logging.info(
                    "Transform (html|4): Pass 4 --> Applying conditional formatting styles..."
                )
                html_with_cf_styles = apply_cf_styles(
                    html_with_updated_links, cf_style_relations
                )

                logging.info(f"Transform (out): Writing output: {dest}")
                output.write(html_with_cf_styles)

            logging.info("Transform: Done!")
            return (True, None)
        except Exception as exc:
            logging.exception("Transform failed for '%s' -> '%s'", source, dest)
            return (False, repr(exc))
        finally:
            if workbook_archive is not None:
                logging.info("Transform (wb|incell): Closing archive...")
                workbook_archive.close()
            if workbook is not None:
                logging.info(f"Transform (wb): Closing wb: {source}")
                workbook.close()

    return transform_xlsx
