from collections.abc import Callable
from typing import TypedDict

from condif2css.css import CssBuilder, CssRulesRegistry
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell, MergedCell
from xx2html.core.types import (
    CellDimensions,
    CellRenderData,
    ColumnRenderData,
    CovaCell,
    WorksheetContents,
)
# from xx2html.core.css import CssRegistry
from xlsx2html.core import (
    rows_from_range,
    format_cell,
    unescape,
    get_cell_id,
    column_index_from_string,
    images_to_data,
    render_attrs,
    render_inline_styles,
)
import logging

COL_WIDTH__FACTOR = 17 / 2.77734375
COL_WIDTH__DEFAULT = 65  # 65px
CELL_HEIGHT__DEFAULT = 19  # 19px
# MAX_COL_REGEX = re.compile(r":(?P<max_col>[a-zA-Z]+)\d+$")
# MAX_COLS_EXCEL = 16_384


def get_worksheet_contents(
    ws: Worksheet,
    # css_registry: CssRegistry,
    # get_css_components_from_cell: Callable[[Cell | CovaCell, dict], Tuple[dict, set]],
    css_rules_registry: CssRulesRegistry,
    css_builder: CssBuilder,
    get_css_from_cell: Callable[[Cell | CovaCell | MergedCell, dict], set[str]],
    locale: None | str = None,
    ws_index: int = -1,
) -> WorksheetContents:
    class VmCellLayoutEntry(TypedDict):
        class_name: str
        vm_id: str
        col_idx_1_based: int
        row_idx_1_based: int
        colspan: int
        rowspan: int

    merged_cell_map = {}
    used_vm_ids: set[str] = set()
    vm_ids_dimension_references: dict[str, CellDimensions] = {}
    vm_cell_vm_ids: dict[str, str] = {}
    vm_cells_layout: list[VmCellLayoutEntry] = []

    merged_cell_ranges = [cell_range.coord for cell_range in ws.merged_cells.ranges]
    excluded_cells = set(
        [
            cell
            for cell_range in merged_cell_ranges
            for rows in rows_from_range(cell_range)
            for cell in rows
        ]
    )

    for cell_range in merged_cell_ranges:
        if ":" not in str(cell_range):
            cell_range_list = list(ws[f"{cell_range}:{cell_range}"])
        else:
            cell_range_list = list(ws[cell_range])

        m_cell = cell_range_list[0][0]

        colspan = len(cell_range_list[0])
        rowspan = len(cell_range_list)
        merged_cell_map[m_cell.coordinate] = {
            "attrs": {
                "colspan": None if colspan <= 1 else colspan,
                "rowspan": None if rowspan <= 1 else rowspan,
            },
            "cells": [c for rows in cell_range_list for c in rows],
        }

        excluded_cells.remove(m_cell.coordinate)

    def get_effective_row_height(row_number: int) -> int:
        row_dim = ws.row_dimensions[row_number]
        if row_dim.hidden:
            return 0
        if row_dim.customHeight and isinstance(row_dim.height, (int, float)):
            return int(round(row_dim.height, 2))
        return CELL_HEIGHT__DEFAULT

    def process_cell(col_idx: int, cell: Cell | CovaCell | MergedCell) -> None:
        if not cell or cell.row is None:
            logging.warning("Cell without row information found, skipping processing.")
            return
        row_dim = ws.row_dimensions[cell.row]

        if cell.coordinate in excluded_cells or row_dim.hidden:
            return

        height = get_effective_row_height(cell.row)

        f_cell = None

        value = cell.value
        if isinstance(value, str):
            value = unescape(value)

        # cell_height_class = css_registry.register_height(height)
        cell_height_class = css_rules_registry.register(
            [css_builder.height(height)]
        )
        classes = set([cell_height_class])
        vm_id = None if not hasattr(cell, "_vm_id") else getattr(cell, "_vm_id")

        cell_data: CellRenderData = {  # initialization of cell_data
            "attrs": {"id": get_cell_id(cell)},
            "column": cell.column,
            "row": cell.row,
            "value": value,
            "formatted_value": format_cell(cell, locale=locale, f_cell=f_cell),
            "style": {},
            "classes": classes,
            "vm_id": vm_id,
        }

        cell_class_name = ""
        if isinstance(vm_id, str):
            cell_class_name = f"cell_{ws_index}_{col_idx}_{row_i}"

            cell_data.update({"formatted_value": ""})
            cell_data["style"].update({"position": "relative", "overflow": "hidden"})
            cell_data["classes"].update(
                [
                    f"vm-richvaluerel_rid{vm_id}",
                    cell_class_name,
                    "incell-image",
                ]
            )

            used_vm_ids.add(vm_id)

        merged_cell_info = merged_cell_map.get(cell.coordinate, {})

        if merged_cell_info:
            cell_data["attrs"].update(  # Update cell_data attrs
                merged_cell_info["attrs"]
            )

        if isinstance(vm_id, str):
            colspan = int(cell_data["attrs"].get("colspan") or 1)
            rowspan = int(cell_data["attrs"].get("rowspan") or 1)
            vm_cells_layout.append(
                {
                    "class_name": cell_class_name,
                    "vm_id": vm_id,
                    "col_idx_1_based": col_idx + 1,
                    "row_idx_1_based": cell.row,
                    "colspan": colspan,
                    "rowspan": rowspan,
                }
            )

        # new_styles, new_classes = get_css_components_from_cell(cell, merged_cell_info)
        # cell_data["style"].update(  # Update cell_data style
        #     new_styles
        # )
        new_classes = get_css_from_cell(cell, merged_cell_info)
        cell_data["classes"].update(  # Update cell_data classes
            new_classes
        )
        data_row.append(cell_data)  # Appending current cell_data to array

    columns_dimensions: dict[str, ColumnRenderData] = {}

    def first_row_process_cell(col_idx: int, cell: Cell | CovaCell | MergedCell) -> None:
        nonlocal columns_dimensions
        column_letter = get_column_letter(col_idx + 1)
        columns_dimensions[column_letter] = {
            "attrs": {},
            "index": column_letter,
            "width": COL_WIDTH__DEFAULT,
            "style": {"visibility": "visible"},
            "hidden": False,
            "collapsed": False,
        }
        process_cell(col_idx, cell)

    current_process_cell = first_row_process_cell

    data_list = []
    for row_i, row in enumerate(ws.iter_rows()):
        data_row = []
        data_list.append(data_row)
        for col_idx, cell in enumerate(row):
            current_process_cell(col_idx, cell)
        current_process_cell = process_cell

    sheet_max_column_index = len(columns_dimensions)

    for _, custom_col_dim in ws.column_dimensions.items():
        if not (
            isinstance(custom_col_dim.min, int) and isinstance(custom_col_dim.max, int)
        ):
            continue

        col_width = int(
            round(
                (custom_col_dim.width * COL_WIDTH__FACTOR)
                if custom_col_dim.customWidth
                else COL_WIDTH__DEFAULT
            )
        )

        for range_col_idx in range(
            custom_col_dim.min, min(custom_col_dim.max, sheet_max_column_index) + 1
        ):
            column_letter = get_column_letter(range_col_idx)
            if column_letter in columns_dimensions:
                column_dimensions = columns_dimensions[column_letter]

                hidden = custom_col_dim.hidden

                column_dimensions["hidden"] = hidden
                column_dimensions["collapsed"] = custom_col_dim.collapsed
                column_dimensions["style"].update(
                    {"visibility": "collapse" if (custom_col_dim.hidden) else "visible"}
                )
                column_dimensions["width"] = col_width if not hidden else 0
            else:
                logging.warning(
                    f"App (Contents): Column '{column_letter}' not found for custom dimensions!"
                )

    col_list: list[ColumnRenderData] = []
    table_width = 0

    for _, col_details in sorted(
        columns_dimensions.items(), key=lambda d: column_index_from_string(d[0])
    ):
        col_list.append(col_details)
        if not col_details["hidden"]:
            table_width += col_details["width"]

    def get_effective_col_width(col_idx_1_based: int) -> int:
        column_letter = get_column_letter(col_idx_1_based)
        col_details = columns_dimensions.get(column_letter)
        if col_details is None:
            return COL_WIDTH__DEFAULT
        width = col_details.get("width")
        if not isinstance(width, int):
            return COL_WIDTH__DEFAULT
        return max(width, 0)

    for vm_cell in vm_cells_layout:
        class_name = vm_cell["class_name"]
        vm_id = vm_cell["vm_id"]
        start_col = vm_cell["col_idx_1_based"]
        start_row = vm_cell["row_idx_1_based"]
        colspan = vm_cell["colspan"]
        rowspan = vm_cell["rowspan"]

        width_px = sum(
            get_effective_col_width(col_idx)
            for col_idx in range(start_col, start_col + colspan)
        )
        height_px = sum(
            get_effective_row_height(row_idx)
            for row_idx in range(start_row, start_row + rowspan)
        )

        if width_px <= 0:
            width_px = COL_WIDTH__DEFAULT
        if height_px <= 0:
            height_px = CELL_HEIGHT__DEFAULT

        vm_ids_dimension_references[class_name] = {
            "width": width_px,
            "height": height_px,
        }
        vm_cell_vm_ids[class_name] = vm_id

    worksheet_contents: WorksheetContents = {
        "rows": data_list,
        "cols": col_list,
        "images": images_to_data(ws),
        "vm_ids": used_vm_ids,
        "vm_ids_dimension_references": vm_ids_dimension_references,
        "vm_cell_vm_ids": vm_cell_vm_ids,
        "table_width": table_width,
    }
    return worksheet_contents


def cova_render_table(
    data: WorksheetContents,  # , append_headers, append_lineno
) -> str:
    html = [
        "".join(
            [
                "<table  ",
                f'style="border:0; border-collapse: collapse; width: {data["table_width"]}px; table-layout: fixed;" '
                if "table_width" in data
                else 'style="border-collapse: collapse; table-layout: fixed;" ',
                " >",
            ]
        )
    ]

    sizes_row = ["<tr>"]

    html.append("<colgroup>")
    for col in data["cols"]:
        html.append(
            '<col {attrs} style="{styles}" data-value="{col_index}">'.format(
                attrs=render_attrs(col.get("attrs")),
                styles=render_inline_styles(col.get("style")),
                col_index=col["index"],
            )
        )
        width = col["width"]
        sizes_row.append(
            '<td style="width: {}px; padding: 0;"></td>'.format(width)
            if isinstance(width, int)
            else "<td></td>"
        )
    html.append("</colgroup>")
    sizes_row.append("</tr>")

    html.append("<tbody>")

    # append_headers(data, html)
    html.append(" ".join(sizes_row))

    for _, row in enumerate(data["rows"]):
        trow = ["<tr>"]
        # append_lineno(trow, i)
        for cell in row:
            images = data["images"].get((cell["column"], cell["row"])) or []
            formatted_images = []

            for img in images:
                styles = render_inline_styles(img["style"])
                img_tag = (
                    '<img width="{width}" height="{height}"'
                    'style="{styles_str}"'
                    'src="{src}"'
                    "/>"
                ).format(styles_str=styles, **img)
                formatted_images.append(img_tag)

            trow.append(
                (
                    '<td {attrs_str} style="{styles_str}" class="{classes_str}">'
                    "{formatted_images}"
                    "{formatted_value}"
                    "{incell_image}"
                    "</td>"
                ).format(
                    attrs_str=render_attrs(cell["attrs"]),
                    styles_str=render_inline_styles(cell["style"]),
                    formatted_images="\n".join(formatted_images),
                    incell_image="<img ></img>" if isinstance(cell["vm_id"], str) else "",
                    classes_str=" ".join(cell["classes"]),
                    **cell,
                )
            )

        trow.append("</tr>")
        html.append("\n".join(trow))

    html.append("</tbody>")
    html.append("</table>")
    return "\n".join(html)
